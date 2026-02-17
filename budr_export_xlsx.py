#!/usr/bin/env python3
"""
BUDR Assessment → XLSX Report Generator
Reads JSON export from AWS Mapper BUDR dashboard and produces
a multi-sheet Excel workbook modeled after CISA ScubaGear reports.

Usage:
    python3 budr_export_xlsx.py budr-assessment.json [-o output.xlsx]

Sheets:
    1. Summary         — tier counts, overall posture, timestamp
    2. Assessments     — one row per resource with RTO/RPO/tier/signals
    3. Findings        — ScubaGear-style: Control ID, Requirement, Result, Criticality, Details
    4. Action Plan     — findings + effort estimate, remediation tracking columns
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Colour palette ──────────────────────────────────────────────
COLORS = {
    'header_bg':   '1B2A4A',   # dark navy
    'header_fg':   'FFFFFF',
    'protected':   'D1FAE5',   # green-100
    'partial':     'FEF3C7',   # amber-100
    'at_risk':     'FEE2E2',   # red-100
    'stripe':      'F8FAFC',   # slate-50
    'border':      'CBD5E1',   # slate-300
    'critical_fg': 'DC2626',
    'high_fg':     'EA580C',
    'medium_fg':   'D97706',
    'low_fg':      '2563EB',
}

TIER_LABELS = {
    'protected': '✓ Protected',
    'partial':   '⚠ Partial',
    'at_risk':   '✗ At Risk',
}

SEVERITY_ORDER = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}

# ScubaGear effort mapping (mirrors _EFFORT_MAP in index.html)
EFFORT_MAP = {
    'BUDR-HA-1': 'moderate', 'BUDR-HA-2': 'moderate', 'BUDR-HA-3': 'quick',
    'BUDR-HA-4': 'moderate', 'BUDR-HA-5': 'moderate', 'BUDR-HA-6': 'quick',
    'BUDR-BAK-1': 'quick',  'BUDR-BAK-2': 'moderate', 'BUDR-BAK-3': 'moderate',
    'BUDR-BAK-4': 'quick',  'BUDR-BAK-5': 'quick',
    'BUDR-DR-1': 'project', 'BUDR-DR-2': 'moderate',
}


# ── Helpers ─────────────────────────────────────────────────────
def style_header(ws, row=1, cols=None):
    """Apply header styling to a row."""
    hdr_font = Font(name='Calibri', bold=True, color=COLORS['header_fg'], size=11)
    hdr_fill = PatternFill('solid', fgColor=COLORS['header_bg'])
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=COLORS['border'])
    hdr_border = Border(bottom=thin)

    for col in range(1, (cols or ws.max_column) + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = hdr_border


def auto_width(ws, min_w=10, max_w=50):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = []
        for cell in col_cells:
            val = str(cell.value) if cell.value else ''
            lengths.append(len(val))
        best = min(max(max(lengths, default=10), min_w), max_w)
        ws.column_dimensions[col_letter].width = best + 2


def tier_fill(tier):
    """Return PatternFill for a tier value."""
    color = COLORS.get(tier, 'FFFFFF')
    return PatternFill('solid', fgColor=color)


def severity_font(sev):
    """Return Font coloured by severity."""
    color = COLORS.get(f'{sev.lower()}_fg', '000000')
    return Font(name='Calibri', size=10, color=color, bold=True)


def stripe_rows(ws, start_row=2):
    """Apply alternating row shading, skip cells that already have a fill."""
    fill = PatternFill('solid', fgColor=COLORS['stripe'])
    no_fill = PatternFill(fill_type=None)
    for idx, row in enumerate(ws.iter_rows(min_row=start_row, max_row=ws.max_row)):
        if idx % 2 == 1:
            for cell in row:
                existing = cell.fill
                has_fill = existing and existing != no_fill and existing.fgColor and str(existing.fgColor.rgb) not in ('00000000', '0')
                if not has_fill:
                    cell.fill = fill


# ── Sheet builders ──────────────────────────────────────────────
def build_summary(wb, data):
    """Sheet 1: Summary overview."""
    ws = wb.create_sheet('Summary', 0)
    summary = data.get('summary', {})
    ts = data.get('timestamp', datetime.now().isoformat())
    total = sum(summary.values())

    # Title block
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = '⛑ BUDR Assessment Report'
    title_cell.font = Font(name='Calibri', bold=True, size=16, color=COLORS['header_bg'])
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws['A3'] = 'Generated'
    ws['B3'] = ts
    ws['A4'] = 'Total Resources'
    ws['B4'] = total
    for r in range(3, 5):
        ws.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws.cell(row=r, column=2).font = Font(size=11)

    # Tier breakdown table
    ws['A6'] = 'Tier'
    ws['B6'] = 'Count'
    ws['C6'] = 'Percentage'
    style_header(ws, row=6, cols=3)

    for i, (tier, label) in enumerate(TIER_LABELS.items()):
        row = 7 + i
        count = summary.get(tier, 0)
        pct = f'{count / total * 100:.1f}%' if total else '0%'
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=pct)
        for c in range(1, 4):
            ws.cell(row=row, column=c).fill = tier_fill(tier)

    # Posture summary
    ws['A11'] = 'Overall Posture'
    ws['A11'].font = Font(bold=True, size=11)
    at_risk = summary.get('at_risk', 0)
    partial = summary.get('partial', 0)
    if at_risk > 0:
        ws['B11'] = f'AT RISK — {at_risk} resource(s) unprotected'
        ws['B11'].font = Font(color=COLORS['critical_fg'], bold=True, size=11)
    elif partial > 0:
        ws['B11'] = f'PARTIAL — {partial} resource(s) need attention'
        ws['B11'].font = Font(color=COLORS['medium_fg'], bold=True, size=11)
    else:
        ws['B11'] = 'PROTECTED — All resources have DR coverage'
        ws['B11'].font = Font(color='059669', bold=True, size=11)

    auto_width(ws)
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 45


def build_assessments(wb, data):
    """Sheet 2: One row per assessed resource."""
    ws = wb.create_sheet('Assessments')
    assessments = data.get('assessments', [])

    headers = ['Type', 'Resource ID', 'Name', 'Tier', 'RTO', 'RPO', 'Signals']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, cols=len(headers))

    for i, a in enumerate(assessments):
        row = i + 2
        profile = a.get('profile') or {}
        tier = profile.get('tier', 'unknown')
        signals = a.get('signals', {})
        sig_str = ', '.join(f'{k}={v}' for k, v in signals.items())

        values = [
            a.get('type', ''),
            a.get('id', ''),
            a.get('name', ''),
            TIER_LABELS.get(tier, tier),
            profile.get('rto', ''),
            profile.get('rpo', ''),
            sig_str,
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            if c == 4:  # Tier column
                cell.fill = tier_fill(tier)

    stripe_rows(ws)
    auto_width(ws)


def build_findings(wb, data):
    """Sheet 3: ScubaGear-style findings (ScubaResults.csv format)."""
    ws = wb.create_sheet('Findings')
    findings = data.get('findings', [])

    # ScubaGear columns
    headers = [
        'Control ID', 'Requirement', 'Result', 'Criticality',
        'Details', 'Resource ID', 'Resource Name', 'Framework',
        'Remediation',
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, cols=len(headers))

    # Sort: CRITICAL first
    sorted_f = sorted(findings, key=lambda f: SEVERITY_ORDER.get(f.get('severity', ''), 99))

    for i, f in enumerate(sorted_f):
        row = i + 2
        sev = f.get('severity', '')
        result = 'Fail' if sev in ('CRITICAL', 'HIGH', 'MEDIUM') else 'Warning'
        values = [
            f.get('control', ''),
            f.get('message', ''),
            result,
            sev,
            f.get('message', ''),
            f.get('resource', ''),
            f.get('resourceName', ''),
            f.get('framework', ''),
            f.get('remediation', ''),
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            if c == 4:  # Criticality column
                cell.font = severity_font(sev)

    stripe_rows(ws)
    auto_width(ws)


def build_action_plan(wb, data):
    """Sheet 4: ScubaGear ActionPlan-style with effort + tracking columns."""
    ws = wb.create_sheet('Action Plan')
    findings = data.get('findings', [])

    headers = [
        'Control ID', 'Requirement', 'Result', 'Criticality',
        'Estimated Effort', 'Resource ID', 'Resource Name',
        'Remediation', 'Non-Compliance Reason',
        'Remediation Owner', 'Target Date', 'Status', 'Justification',
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header(ws, cols=len(headers))

    sorted_f = sorted(findings, key=lambda f: SEVERITY_ORDER.get(f.get('severity', ''), 99))

    for i, f in enumerate(sorted_f):
        row = i + 2
        sev = f.get('severity', '')
        ctrl = f.get('control', '')
        effort = EFFORT_MAP.get(ctrl, 'moderate')
        result = 'Fail' if sev in ('CRITICAL', 'HIGH', 'MEDIUM') else 'Warning'
        values = [
            ctrl,
            f.get('message', ''),
            result,
            sev,
            effort.title(),
            f.get('resource', ''),
            f.get('resourceName', ''),
            f.get('remediation', ''),
            f.get('message', ''),       # Non-Compliance Reason
            '',                          # Owner (user fills)
            '',                          # Target Date (user fills)
            'Open',                      # Status
            '',                          # Justification (user fills)
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            if c == 4:
                cell.font = severity_font(sev)
            if c == 5:  # Effort column colour
                effort_colors = {'Quick': 'D1FAE5', 'Moderate': 'FEF3C7', 'Project': 'FEE2E2'}
                cell.fill = PatternFill('solid', fgColor=effort_colors.get(effort.title(), 'FFFFFF'))

    stripe_rows(ws)
    auto_width(ws)
    # Widen tracking columns
    for col_letter in ['J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 18


# ── Main ────────────────────────────────────────────────────────
def generate_report(json_path, output_path=None):
    """Read BUDR JSON and write XLSX report."""
    with open(json_path) as f:
        data = json.load(f)

    if not output_path:
        stem = Path(json_path).stem
        output_path = str(Path(json_path).parent / f'{stem}.xlsx')

    wb = pd.ExcelWriter(output_path, engine='openpyxl')
    # Create workbook via openpyxl directly (pandas just for the writer context)
    workbook = wb.book

    build_summary(workbook, data)
    build_assessments(workbook, data)
    build_findings(workbook, data)
    build_action_plan(workbook, data)

    # Remove default sheet if empty
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    wb.close()
    print(f'✓ Report saved: {output_path}')
    print(f'  Sheets: {", ".join(workbook.sheetnames)}')
    a = data.get('assessments', [])
    f = data.get('findings', [])
    print(f'  {len(a)} assessments, {len(f)} findings')
    return output_path


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='BUDR Assessment → XLSX Report')
    parser.add_argument('json_file', help='Path to budr-assessment.json')
    parser.add_argument('-o', '--output', help='Output .xlsx path (default: same name as input)')
    args = parser.parse_args()
    generate_report(args.json_file, args.output)
